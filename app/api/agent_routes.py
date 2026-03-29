"""
app/api/agent_routes.py
────────────────────────
New agent-pipeline endpoint that replaces the old monolithic upload flow.
Also provides /documents list & detail, /search, /analytics, and /export.
"""
from __future__ import annotations

import csv
import io
import json
import time
from typing import Any, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse

from agents.orchestrator import AgentOrchestrator
from app.auth import get_current_user
from app.database import (
    get_dashboard_stats,
    get_document,
    list_documents,
    record_event,
    save_document,
    search_documents,
)
from app.models.schemas import (
    AskQuestionRequest,
    AskQuestionResponse,
    AnalyticsData,
    AnalyticsEvent,
)
from utils.logger import logger

router = APIRouter(prefix="/api/v2", tags=["Enterprise Agent API"])

# Single orchestrator instance (thread-safe for reads)
_orchestrator = AgentOrchestrator()

_ALLOWED_TYPES = {
    "application/pdf", "image/jpeg", "image/png",
    "image/tiff", "image/bmp", "image/webp",
}
_MAX_FILE_SIZE = 20 * 1024 * 1024


# ─── Upload (full agent pipeline) ─────────────────────────────────────────────

@router.post("/upload", summary="Upload document through full agent pipeline")
async def upload_document_agent(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
) -> JSONResponse:
    content = await file.read()

    if len(content) > _MAX_FILE_SIZE:
        raise HTTPException(413, f"File exceeds 20 MB limit.")
    if file.content_type not in _ALLOWED_TYPES:
        raise HTTPException(415, f"Unsupported type: {file.content_type}")

    t0 = time.perf_counter()
    try:
        ctx = _orchestrator.process_document(file.filename or "document", content)
    except RuntimeError as exc:
        raise HTTPException(500, str(exc)) from exc
    except Exception as exc:
        logger.exception("Agent pipeline failed.")
        raise HTTPException(500, "Internal server error.") from exc

    # Persist to DB in background
    background_tasks.add_task(save_document, ctx, current_user.get("sub"))
    background_tasks.add_task(
        record_event,
        "upload",
        ctx.get("document_id"),
        str(ctx.get("doc_type", "unknown")),
        ctx.get("pipeline_elapsed_ms", 0.0),
        ctx.get("classification_confidence", 0.0),
        current_user.get("sub"),  # user_id
    )

    fields = ctx.get("extracted_fields")
    fields_dict: dict = {}
    if fields is not None:
        fields_dict = fields.model_dump() if hasattr(fields, "model_dump") else dict(fields)

    return JSONResponse({
        "document_id":              ctx.get("document_id"),
        "filename":                 ctx.get("filename"),
        "document_type":            str(ctx.get("doc_type", "unknown")).split(".")[-1].lower(),
        "language":                 str(ctx.get("language", "unknown")).split(".")[-1].lower(),
        "page_count":               ctx.get("page_count", 1),
        "character_count":          len(ctx.get("clean_text", "")),
        "raw_text_preview":         ctx.get("clean_text", "")[:500],
        "ocr_confidence":           ctx.get("ocr_confidence", 0.0),
        "classification_confidence": ctx.get("classification_confidence", 0.0),
        "extraction_confidence":    ctx.get("extraction_confidence", 0.0),
        "fields":                   fields_dict,
        "validation_report":        ctx.get("validation_report", {}),
        "workflow_actions":         ctx.get("workflow_actions", []),
        "layout_info":              ctx.get("layout_info", {}),
        "pipeline_elapsed_ms":      ctx.get("pipeline_elapsed_ms", 0.0),
        "status":                   ctx.get("status", "pending"),
        "message":                  "Document processed successfully through the agent pipeline.",
    })


# ─── Q&A ──────────────────────────────────────────────────────────────────────

@router.post("/ask", response_model=AskQuestionResponse, summary="RAG question answering")
async def ask_question_agent(body: AskQuestionRequest) -> AskQuestionResponse:
    try:
        result = _orchestrator.answer_question(body.document_id, body.question)
        return AskQuestionResponse(
            document_id=body.document_id,
            question=body.question,
            answer=result["answer"],
            source_chunks=result["source_chunks"],
        )
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(500, str(exc)) from exc


# ─── Document list & detail ────────────────────────────────────────────────────

@router.get("/documents", summary="List all processed documents (user-isolated with admin override)")
async def list_docs(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    current_user: dict = Depends(get_current_user),
) -> JSONResponse:
    user_id = current_user.get("sub")
    user_role = current_user.get("role", "user")
    
    # Admin users see ALL documents, regular users see only their own
    docs = list_documents(limit=limit, offset=offset, user_id=user_id, user_role=user_role)
    # Standardize keys for frontend compatibility
    standardized = []
    for d in docs:
        standardized.append({
            "document_id":              d["document_id"],
            "filename":                 d["filename"],
            "document_type":            str(d["doc_type"]).split(".")[-1].lower(),
            "language":                 str(d["language"]).split(".")[-1].lower(),
            "page_count":               d["page_count"],
            "character_count":          d["character_count"],
            "ocr_confidence":           d["ocr_confidence"],
            "classification_confidence": d["classification_conf"],
            "extraction_confidence":    d["extraction_conf"],
            "validation_passed":        bool(d["validation_passed"]),
            "validation_score":         d["validation_score"],
            "pipeline_elapsed_ms":      d["pipeline_elapsed_ms"],
            "created_at":               d["created_at"],
            "fields":                   json.loads(d["fields_json"]) if d.get("fields_json") else {},
            "validation_report":        json.loads(d["validation_json"]) if d.get("validation_json") else {},
            "layout_info":              json.loads(d["layout_json"]) if d.get("layout_json") else {},
            "status":                   d.get("status", "pending"),
        })
    return JSONResponse({"total": len(docs), "documents": standardized})


@router.get("/documents/{document_id}", summary="Get document metadata")
async def get_doc(document_id: str) -> JSONResponse:
    doc = get_document(document_id)
    if not doc:
        raise HTTPException(404, f"Document '{document_id}' not found.")
    
    # Standardize keys for frontend compatibility
    # Standardize keys for frontend compatibility
    return JSONResponse({
        "document_id":              doc["document_id"],
        "filename":                 doc["filename"],
        "document_type":            str(doc["doc_type"]).split(".")[-1].lower(),
        "language":                 str(doc["language"]).split(".")[-1].lower(),
        "page_count":               doc["page_count"],
        "character_count":          doc["character_count"],
        "ocr_confidence":           doc["ocr_confidence"],
        "classification_confidence": doc["classification_conf"],
        "extraction_confidence":    doc["extraction_conf"],
        "validation_passed":        bool(doc["validation_passed"]),
        "validation_score":         doc["validation_score"],
        "pipeline_elapsed_ms":      doc["pipeline_elapsed_ms"],
        "created_at":               doc["created_at"],
        "fields":                   json.loads(doc["fields_json"]) if doc.get("fields_json") else {},
        "validation_report":        json.loads(doc["validation_json"]) if doc.get("validation_json") else {},
        "layout_info":              json.loads(doc["layout_json"]) if doc.get("layout_json") else {},
        "status":                   doc.get("status", "pending"),
    })


@router.patch("/documents/{document_id}/status", summary="Update document approval status")
async def update_doc_status(document_id: str, body: dict) -> JSONResponse:
    new_status = body.get("status")
    if not new_status:
        raise HTTPException(400, "Missing 'status' in request body.")
    
    from app.database import update_document_status
    try:
        update_document_status(document_id, new_status)
        return JSONResponse({
            "document_id": document_id,
            "status": new_status,
            "message": f"Document status updated to '{new_status}'."
        })
    except Exception as exc:
        raise HTTPException(500, str(exc)) from exc


# ─── Search ───────────────────────────────────────────────────────────────────

@router.get("/search", summary="Keyword + filter document search")
async def search_docs(
    q: str = Query("", description="Keyword query"),
    doc_type: Optional[str] = Query(None, description="Filter by document type"),
    limit: int = Query(20, ge=1, le=100),
) -> JSONResponse:
    results = search_documents(query=q, doc_type=doc_type, limit=limit)
    standardized = []
    for d in results:
        standardized.append({
            "document_id":              d["document_id"],
            "filename":                 d["filename"],
            "document_type":            str(d["doc_type"]).split(".")[-1].lower(),
            "language":                 str(d["language"]).split(".")[-1].lower(),
            "page_count":               d["page_count"],
            "character_count":          d["character_count"],
            "ocr_confidence":           d["ocr_confidence"],
            "classification_confidence": d["classification_conf"],
            "extraction_confidence":    d["extraction_conf"],
            "validation_passed":        bool(d["validation_passed"]),
            "pipeline_elapsed_ms":      d["pipeline_elapsed_ms"],
            "created_at":               d["created_at"],
            "fields":                   json.loads(d["fields_json"]) if d.get("fields_json") else {},
            "validation_report":        json.loads(d["validation_json"]) if d.get("validation_json") else {},
            "layout_info":              json.loads(d["layout_json"]) if d.get("layout_json") else {},
        })
    return JSONResponse({"query": q, "total": len(results), "results": standardized})


# ─── Analytics ────────────────────────────────────────────────────────────────

@router.get("/analytics", summary="Role-based aggregated processing analytics", response_model=AnalyticsData)
async def analytics(current_user: dict = Depends(get_current_user)) -> AnalyticsData:
    user_id = current_user.get("sub")
    user_role = current_user.get("role", "user")
    
    # Admin users see global stats, regular users see only their own
    stats = get_dashboard_stats(user_id=user_id, user_role=user_role)
    events = [AnalyticsEvent(**e) for e in stats["recent_events"]]
    
    # Calculate percentage trends
    today_trend = _calculate_percentage_change(stats["today_count"], stats["yesterday_count"])
    weekly_trend = _calculate_percentage_change(stats["this_week_count"], stats["last_week_count"])
    success_rate_trend = _calculate_percentage_change(
        stats["current_success_rate"], 
        stats["previous_success_rate"]
    )
    processing_time_trend = _calculate_percentage_change(
        stats["current_avg_processing_ms"],
        stats["previous_avg_processing_ms"]
    )
    
    return AnalyticsData(
        total_documents=stats["total_documents"],
        total_events=stats["total_events"],
        by_type=stats["by_type"],
        avg_processing_ms=stats["avg_processing_ms"],
        avg_confidence=stats["avg_confidence"],
        recent_events=events,
        today_trend=today_trend,
        weekly_trend=weekly_trend,
        success_rate_trend=success_rate_trend,
        processing_time_trend=processing_time_trend
    )


def _calculate_percentage_change(current: float | int, previous: float | int) -> float | None:
    """Calculate percentage change between two values.
    
    Args:
        current: Current period value
        previous: Previous period value
    
    Returns:
        Percentage change (positive or negative), or None if previous is 0
    """
    if previous == 0:
        return None if current == 0 else 100.0  # If previous was 0 and current > 0, show 100% growth
    return round(((current - previous) / previous) * 100, 1)


# ─── Agent pipeline status ────────────────────────────────────────────────────

@router.get("/agents/status", summary="List all registered agents")
async def agent_status() -> JSONResponse:
    return JSONResponse(_orchestrator.get_pipeline_status())


@router.get("/documents/{document_id}/preview", summary="Get document preview image")
async def get_document_preview(document_id: str) -> FileResponse:
    doc = get_document(document_id)
    if not doc:
        raise HTTPException(404, f"Document '{document_id}' not found.")
    
    from pathlib import Path
    from utils.config import get_settings
    _settings = get_settings()
    
    doc_dir = Path(_settings.upload_dir) / document_id
    preview_path = doc_dir / "preview.png"
    
    # 1. Check for dedicated preview image
    if preview_path.exists():
        return FileResponse(preview_path)
    
    # 2. Check original file and convert if needed
    filename = doc["filename"]
    original_path = doc_dir / filename
    if not original_path.exists():
         raise HTTPException(404, detail="Source file not found")
         
    ext = original_path.suffix.lower()
    
    # Browser-native images can be served directly
    if ext in {".jpg", ".jpeg", ".png", ".webp"}:
        return FileResponse(original_path)
        
    # PDF conversion fallback
    if ext == ".pdf":
         try:
             from ocr.ocr_service import _pdf_to_images
             content = original_path.read_bytes()
             images = _pdf_to_images(content)
             if images:
                 preview_path.parent.mkdir(parents=True, exist_ok=True)
                 images[0].save(preview_path, format="PNG")
                 return FileResponse(preview_path)
         except Exception as exc:
             logger.warning("V2 PDF preview generation failed: %s", exc)

    raise HTTPException(status_code=404, detail="Preview not available")


@router.post("/documents/{document_id}/reextract", summary="Re-run extraction for an existing document")
async def reextract_document(document_id: str) -> JSONResponse:
    doc = get_document(document_id)
    if not doc:
        raise HTTPException(404, f"Document '{document_id}' not found.")
    
    from pathlib import Path
    from utils.config import get_settings
    from app.models.schemas import DocumentType

    _settings = get_settings()
    doc_dir = Path(_settings.upload_dir) / document_id
    text_path = doc_dir / "ocr_text.txt"
    
    clean_text = ""
    if text_path.exists():
        clean_text = text_path.read_text(encoding="utf-8")
    else:
        # Fallback: re-run OCR if original file exists
        original_path = doc_dir / doc["filename"]
        if original_path.exists():
            from ocr.ocr_service import extract_text
            ocr_res = extract_text(original_path.read_bytes(), doc["filename"])
            clean_text = ocr_res["raw_text"]
            # Save it for next time
            text_path.write_text(clean_text, encoding="utf-8")
        else:
            raise HTTPException(404, "Original document text/file not found for re-extraction.")

    # Normalize doc_type for extractor
    db_type = doc["doc_type"]
    try:
        if "." in db_type: db_type = db_type.split(".")[-1]
        doc_type_enum = DocumentType(db_type.lower())
    except:
        doc_type_enum = DocumentType.unknown

    # Run re-extraction via orchestrator
    res_ctx = _orchestrator.reextract(document_id, clean_text, doc_type_enum)
    
    # Standardize result for DB saving (needs clean_text for DB helpers if they use it)
    res_ctx["clean_text"] = clean_text
    res_ctx["filename"] = doc["filename"]
    res_ctx["doc_type"] = doc_type_enum
    res_ctx["language"] = doc["language"]
    
    # Save back to DB
    save_document(res_ctx)
    
    fields = res_ctx.get("extracted_fields")
    fields_dict = fields.model_dump() if hasattr(fields, "model_dump") else (fields or {})

    return JSONResponse({
        "document_id": document_id,
        "fields": fields_dict,
        "extraction_confidence": res_ctx.get("extraction_confidence", 0.0),
        "validation_report": res_ctx.get("validation_report", {}),
        "message": "Re-extraction completed successfully."
    })


# ─── Export ───────────────────────────────────────────────────────────────────

@router.get("/export/{document_id}", summary="Export extracted fields")
async def export_document(
    document_id: str,
    format: str = Query("json", description="json | csv | excel"),
) -> Any:
    doc = get_document(document_id)
    if not doc:
        raise HTTPException(404, f"Document '{document_id}' not found.")

    # Build flat export dict from full extracted fields 
    fields = json.loads(doc["fields_json"]) if doc.get("fields_json") else {}
    export_data = {k: v for k, v in fields.items() if v is not None}
    # Add metadata
    export_data["document_id"] = doc["document_id"]
    export_data["filename"] = doc["filename"]
    export_data["doc_type"] = str(doc["doc_type"]).split(".")[-1].lower()

    if format == "json":
        return StreamingResponse(
            io.BytesIO(json.dumps(export_data, indent=2, ensure_ascii=False).encode()),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{document_id}.json"'},
        )

    elif format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=list(export_data.keys()))
        writer.writeheader()
        writer.writerow(export_data)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{document_id}.csv"'},
        )

    elif format == "excel":
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Fields"
        ws.append(list(export_data.keys()))
        ws.append([str(v) if v is not None else "" for v in export_data.values()])

        # Style header row
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill("solid", fgColor="4338CA")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{document_id}.xlsx"'},
        )

    else:
        raise HTTPException(400, f"Unknown format '{format}'. Use json, csv, or excel.")
